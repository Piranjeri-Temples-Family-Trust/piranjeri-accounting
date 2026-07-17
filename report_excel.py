"""
report_excel.py — Excel report generators for Piranjeri Temples Family Trust
Uses openpyxl. Each function accepts pre-fetched data (no DB calls) and returns xlsx bytes.

Functions:
    trial_balance_xlsx(fy, df)
    ie_statement_xlsx(fy, year_end_str, data)
    balance_sheet_xlsx(fy, data)
    festival_pnl_xlsx(fy, rows)
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette (hex, no #) ───────────────────────────────────────────────────────
BLUE_HEX   = "1A3C6E"
INDIGO_HEX = "4A5299"
LGREY_HEX  = "F2F2F2"
MGREY_HEX  = "CCCCCC"
RED_HEX    = "CC0000"
GREEN_HEX  = "006600"
WHITE_HEX  = "FFFFFF"

MONEY  = '#,##0.00'


# ── Style factories ───────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(bold=False, size=10, color=None, italic=False):
    return Font(name="Calibri", bold=bold, italic=italic, size=size,
                color=color or "000000")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin(color=MGREY_HEX):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_top(color=BLUE_HEX, style="medium"):
    return Border(top=Side(style=style, color=color))

def _border_tb(color=BLUE_HEX, style="medium"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s)

def _set_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _trust_header(ws, title, subtitle, fy, ncols):
    """Write rows 1-4 trust header. Returns next available row (5)."""
    for row_idx, (text, size, bold, color) in enumerate([
        ("Piranjeri Temples Family Trust", 14, True,  BLUE_HEX),
        (title,                            11, True,  INDIGO_HEX),
        (subtitle,                          9, False, "555555"),
        (f"Financial Year {fy}",            9, False, "888888"),
    ], start=1):
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx,   end_column=ncols)
        c = ws.cell(row_idx, 1, text)
        c.font      = _font(bold=bold, size=size, color=color)
        c.alignment = _align("center")
        ws.row_dimensions[row_idx].height = {1: 22, 2: 18, 3: 14, 4: 13}[row_idx]
    return 5   # first data row


def _write_col_hdr(ws, row, headers, hdr_cols=None, money_cols=None):
    """Write styled column-header row. hdr_cols: which cols get money alignment."""
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row, col, hdr)
        c.font      = _font(bold=True, size=9, color=WHITE_HEX)
        c.fill      = _fill(BLUE_HEX)
        c.border    = _border_thin()
        c.alignment = _align("right" if (money_cols and col in money_cols) else "center")
    ws.row_dimensions[row].height = 18


def _make_xlsx(wb):
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# 1. TRIAL BALANCE
# ══════════════════════════════════════════════════════════════════════════════

def trial_balance_xlsx(fy, df):
    """
    fy : str  e.g. '2025-26'
    df : pandas DataFrame with columns: code, name, account_type, dr_balance, cr_balance
    Returns xlsx bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Trial Balance {fy}"

    r = _trust_header(ws, "Trial Balance",
                      f"As at 31 March 20{fy.split('-')[1]}", fy, 4)

    _write_col_hdr(ws, r, ["Code", "Account", "Debit (₹)", "Credit (₹)"],
                   money_cols=[3, 4])
    r += 1

    GROUP_LABELS = {
        "ASSET": "Assets", "INCOME": "Income", "EXPENDITURE": "Expenditure",
        "FUND": "Funds", "LIABILITY": "Liabilities",
    }
    GROUP_ORDER = ["ASSET", "INCOME", "EXPENDITURE", "FUND", "LIABILITY"]

    total_dr = 0.0
    total_cr = 0.0
    stripe   = False

    for gtype in GROUP_ORDER:
        gdf = df[df["account_type"] == gtype]
        if gdf.empty:
            continue

        # Group header row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1, GROUP_LABELS.get(gtype, gtype))
        c.font      = _font(bold=True, size=9, color=WHITE_HEX)
        c.fill      = _fill(INDIGO_HEX)
        c.alignment = _align()
        ws.row_dimensions[r].height = 14
        r += 1

        for _, row in gdf.iterrows():
            dr = float(row.get("dr_balance", 0) or 0)
            cr = float(row.get("cr_balance", 0) or 0)
            total_dr += dr
            total_cr += cr
            bg = _fill(LGREY_HEX) if stripe else _fill(WHITE_HEX)
            stripe = not stripe

            for col, val, align in [
                (1, str(row["code"]), "left"),
                (2, str(row["name"]), "left"),
                (3, dr if dr > 0.005 else None, "right"),
                (4, cr if cr > 0.005 else None, "right"),
            ]:
                c = ws.cell(r, col, val)
                c.font      = _font(size=9)
                c.fill      = bg
                c.border    = _border_thin()
                c.alignment = _align(align)
                if col in (3, 4) and val is not None:
                    c.number_format = MONEY
            ws.row_dimensions[r].height = 14
            r += 1

    # Total row
    tb = _border_tb(BLUE_HEX, "medium")
    ws.cell(r, 2, "TOTAL").font  = _font(bold=True, size=9)
    ws.cell(r, 2).border = tb

    for col, val in [(3, total_dr), (4, total_cr)]:
        c = ws.cell(r, col, val)
        c.font          = _font(bold=True, size=9)
        c.number_format = MONEY
        c.alignment     = _align("right")
        c.border        = tb

    _set_width(ws, 1,  9)
    _set_width(ws, 2, 40)
    _set_width(ws, 3, 18)
    _set_width(ws, 4, 18)

    return _make_xlsx(wb)


# ══════════════════════════════════════════════════════════════════════════════
# 2. I&E STATEMENT
# ══════════════════════════════════════════════════════════════════════════════

def ie_statement_xlsx(fy, year_end_str, data):
    """
    fy          : str   e.g. '2025-26'
    year_end_str: str   e.g. '31 March 2026'
    data        : dict  same keys as ie_statement_pdf
    Returns xlsx bytes.
    """
    d = data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"IE {fy}"

    # 6 columns: ExpDesc | ExpInner | ExpOuter | IncDesc | IncInner | IncOuter
    ncols = 6
    r = _trust_header(ws, "Income & Expenditure Account",
                      f"For the year ended {year_end_str}", fy, ncols)

    _write_col_hdr(ws, r,
                   ["Expenditure", "", "Amount (₹)", "Income", "", "Amount (₹)"],
                   money_cols=[3, 6])
    r += 1

    ie_result    = d["ie_result"]
    grand_total  = d["grand_total"]
    slabel       = d["surplus_label"]
    bal_amt      = d["balancing_amt"]

    festival_exp    = d["e01"]+d["e04"]+d["e02"]+d["e05"]+d["e06"]+d["e03"]
    other_exp       = d["e08"] + d["e09"]
    total_donations = d["i01"]+d["i_aadi"]+d["i02"]+d["i03"]+d["i04"]+d["i05"]
    total_interest  = d["i_sb"] + d["i_fd"]

    SECT_FILL = _fill("E8ECF5")
    LGREY     = _fill(LGREY_HEX)
    WHITE     = _fill(WHITE_HEX)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def write_cell(row, col, val, bold=False, money=False,
                   fill=None, border=None, align="left", color=None):
        c = ws.cell(row, col, val)
        c.font      = _font(bold=bold, size=9, color=color)
        c.alignment = _align(align)
        if money and isinstance(val, (int, float)):
            c.number_format = MONEY
        if fill:
            c.fill = fill
        if border:
            c.border = border
        ws.row_dimensions[row].height = 13

    def section_hdr(row, e_label, i_label):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        for col, lbl in [(1, e_label), (4, i_label)]:
            c = ws.cell(row, col, lbl)
            c.font      = _font(bold=True, italic=True, size=9, color="444444")
            c.fill      = SECT_FILL
            c.alignment = _align()
        ws.row_dimensions[row].height = 13

    def item_row(row, e_label, e_amt, i_label, i_amt, stripe=False):
        bg = LGREY if stripe else WHITE
        write_cell(row, 1, e_label, fill=bg)
        write_cell(row, 2, e_amt,  fill=bg, money=True, align="right")
        write_cell(row, 3, None,   fill=bg)
        write_cell(row, 4, i_label, fill=bg)
        write_cell(row, 5, i_amt,  fill=bg, money=True, align="right")
        write_cell(row, 6, None,   fill=bg)

    def subtotal_row(row, e_label, e_val, i_label, i_val):
        top = _border_top(MGREY_HEX, "thin")
        for col, lbl, val in [(1, e_label, e_val), (4, i_label, i_val)]:
            c = ws.cell(row, col, lbl)
            c.font   = _font(bold=True, size=9)
            c.border = top
            c.alignment = _align()
        for col, val in [(3, e_val), (6, i_val)]:
            c = ws.cell(row, col, val)
            c.font          = _font(bold=True, size=9)
            c.number_format = MONEY
            c.alignment     = _align("right")
            c.border        = top
        ws.row_dimensions[row].height = 13

    def total_row(row, e_label, e_val, i_label, i_val):
        tb = _border_tb(BLUE_HEX, "medium")
        for col, lbl, val in [(1, e_label, e_val), (4, i_label, i_val)]:
            c = ws.cell(row, col, lbl)
            c.font   = _font(bold=True, size=9)
            c.border = tb
            c.alignment = _align()
        for col, val in [(3, e_val), (6, i_val)]:
            c = ws.cell(row, col, val)
            c.font          = _font(bold=True, size=9)
            c.number_format = MONEY
            c.alignment     = _align("right")
            c.border        = tb
        ws.row_dimensions[row].height = 13

    def blank_row(row):
        ws.row_dimensions[row].height = 6

    # ── Section 1: Festival income/expenditure ────────────────────────────────
    section_hdr(r, "Expenditure towards", "Donations received for"); r += 1

    exp_items = [
        ("  Nithya Pooja",     d["e01"]),
        ("  Aadi Pooram",      d["e04"]),
        ("  Pradosham",        d["e02"]),
        ("  Garuda Seva",      d["e05"]),
        ("  Varushabhishekam", d["e06"]),
        ("  Panguni Uthram",   d["e03"]),
    ]
    inc_items = [
        ("  Nithya Pooja",     d["i01"]),
        ("  Aadi Pooram",      d["i_aadi"]),
        ("  Pradosham",        d["i02"]),
        ("  Garuda Seva",      d["i03"]),
        ("  Varushabhishekam", d["i04"]),
        ("  Panguni Uthram",   d["i05"]),
    ]
    # Filter zero items (align exp and inc separately — pad shorter)
    exp_items = [(l, v) for l, v in exp_items if v > 0.005]
    inc_items = [(l, v) for l, v in inc_items if v > 0.005]
    maxlen = max(len(exp_items), len(inc_items))
    while len(exp_items) < maxlen: exp_items.append(("", None))
    while len(inc_items) < maxlen: inc_items.append(("", None))

    for i, ((el, ev), (il, iv)) in enumerate(zip(exp_items, inc_items)):
        item_row(r, el, ev, il, iv, stripe=(i % 2 == 1)); r += 1

    subtotal_row(r, "Sub-total", festival_exp, "Sub-total", total_donations); r += 1
    blank_row(r); r += 1

    # ── Section 2: Other exp / Interest income ────────────────────────────────
    section_hdr(r, "Other Expenses", "Interest Income"); r += 1

    other_pairs = [
        (("  Bank Charges", d["e08"] if d["e08"] > 0.005 else None),
         ("  Interest on Savings Bank",   d["i_sb"] if d["i_sb"] > 0.005 else None)),
        (("  Audit Fees",   d["e09"] if d["e09"] > 0.005 else None),
         ("  Interest on Fixed Deposits", d["i_fd"] if d["i_fd"] > 0.005 else None)),
    ]
    for i, ((el, ev), (il, iv)) in enumerate(other_pairs):
        if ev is not None or iv is not None:
            item_row(r, el or "", ev, il or "", iv, stripe=(i % 2 == 1)); r += 1

    subtotal_row(r, "Sub-total", other_exp, "Sub-total", total_interest); r += 1
    blank_row(r); r += 1

    # ── Balancing figure ──────────────────────────────────────────────────────
    if ie_result >= 0:
        # Surplus → show on expenditure side
        write_cell(r, 1, slabel, bold=True)
        c = ws.cell(r, 3, bal_amt); c.font = _font(bold=True, size=9)
        c.number_format = MONEY; c.alignment = _align("right")
        blank_row(r + 1)
        r += 2
    else:
        # Deficit → show on income side
        write_cell(r, 4, slabel, bold=True)
        c = ws.cell(r, 6, bal_amt); c.font = _font(bold=True, size=9)
        c.number_format = MONEY; c.alignment = _align("right")
        blank_row(r + 1)
        r += 2

    total_row(r, "TOTAL", grand_total, "TOTAL", grand_total)

    # Vertical divider between Exp and Inc halves
    for row_idx in range(5, r + 1):
        c = ws.cell(row_idx, 3)
        existing = c.border
        c.border = Border(
            left=existing.left, top=existing.top, bottom=existing.bottom,
            right=Side(style="medium", color=BLUE_HEX)
        )

    _set_width(ws, 1, 30); _set_width(ws, 2, 14); _set_width(ws, 3, 16)
    _set_width(ws, 4, 30); _set_width(ws, 5, 14); _set_width(ws, 6, 16)

    return _make_xlsx(wb)


# ══════════════════════════════════════════════════════════════════════════════
# 3. BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════

def balance_sheet_xlsx(fy, data):
    """
    fy   : str  e.g. '2025-26'
    data : dict keys: a01,a02,a03,a04,a05,l01,l02,l03,l04,l05,
                       ob_l01,ob_l02,ob_l03,i06_cr,e07_dr,total_assets,total_fl
    Returns xlsx bytes.
    """
    d = data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Balance Sheet {fy}"
    year_end = fy.split("-")[1]

    r = _trust_header(ws, "Balance Sheet",
                      f"As at 31 March 20{year_end}", fy, 4)

    _write_col_hdr(ws, r,
                   ["Funds & Liabilities", "Amount (₹)", "Assets", "Amount (₹)"],
                   money_cols=[2, 4])
    r += 1

    a01 = d["a01"]; a02 = d["a02"]; a03 = d["a03"]
    a04 = d["a04"]; a05 = d["a05"]
    l01 = d["l01"]; l02 = d["l02"]; l03 = d["l03"]
    l04 = d["l04"]; l05 = d["l05"]
    ob_l01 = d["ob_l01"]; ob_l02 = d["ob_l02"]; ob_l03 = d["ob_l03"]
    i06_cr = d["i06_cr"]; e07_dr = d["e07_dr"]
    total_assets = d["total_assets"]; total_fl = d["total_fl"]
    l03_label    = "Non-Corpus Fund (Deficit)" if l03 < 0 else "Non-Corpus Fund"
    l03_movement = l03 - ob_l03

    # Helper: write left/right cell pair
    def wl(row, label, amount=None, bold=False):
        c = ws.cell(row, 1, label)
        c.font = _font(bold=bold, size=9)
        c.alignment = _align()
        c.border = _border_thin()
        if amount is not None:
            ca = ws.cell(row, 2, amount)
            ca.font = _font(bold=bold, size=9)
            ca.number_format = MONEY
            ca.alignment = _align("right")
            ca.border = _border_thin()
        ws.row_dimensions[row].height = 14

    def wr(row, label, amount=None, bold=False):
        c = ws.cell(row, 3, label)
        c.font = _font(bold=bold, size=9)
        c.alignment = _align()
        c.border = _border_thin()
        if amount is not None:
            ca = ws.cell(row, 4, amount)
            ca.font = _font(bold=bold, size=9)
            ca.number_format = MONEY
            ca.alignment = _align("right")
            ca.border = _border_thin()
        ws.row_dimensions[row].height = 14

    def hdr_l(row, label):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1, label)
        c.font = _font(bold=True, size=9, color=WHITE_HEX)
        c.fill = _fill(BLUE_HEX)
        c.alignment = _align()
        ws.row_dimensions[row].height = 14

    def hdr_r(row, label):
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c = ws.cell(row, 3, label)
        c.font = _font(bold=True, size=9, color=WHITE_HEX)
        c.fill = _fill(BLUE_HEX)
        c.alignment = _align()
        ws.row_dimensions[row].height = 14

    def total_l(row, label, amount):
        tb = _border_tb(BLUE_HEX, "medium")
        c = ws.cell(row, 1, label); c.font = _font(bold=True, size=9); c.border = tb
        c.alignment = _align()
        ca = ws.cell(row, 2, amount)
        ca.font = _font(bold=True, size=9); ca.number_format = MONEY
        ca.alignment = _align("right"); ca.border = tb
        ws.row_dimensions[row].height = 14

    def total_r(row, label, amount):
        tb = _border_tb(BLUE_HEX, "medium")
        c = ws.cell(row, 3, label); c.font = _font(bold=True, size=9); c.border = tb
        c.alignment = _align()
        ca = ws.cell(row, 4, amount)
        ca.font = _font(bold=True, size=9); ca.number_format = MONEY
        ca.alignment = _align("right"); ca.border = tb
        ws.row_dimensions[row].height = 14

    start_r = r

    # LEFT SIDE — Funds & Liabilities
    hdr_l(r, "Funds & Liabilities"); r += 1

    wl(r, "Corpus Fund", bold=True); r += 1
    wl(r, f"  Opening balance"); ws.cell(r, 2, ob_l01); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    wl(r, "  Additions: Nil"); r += 1
    wl(r, "Corpus Fund Total", l01, bold=True); r += 1
    ws.row_dimensions[r].height = 5; r += 1

    wl(r, "Renovation Fund", bold=True); r += 1
    wl(r, f"  Opening balance"); ws.cell(r,2,ob_l02); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    wl(r, "  Add: Renovation income"); ws.cell(r,2,i06_cr); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    wl(r, f"  Less: Renovation exp"); ws.cell(r,2,e07_dr); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    wl(r, "Renovation Fund Total", l02, bold=True); r += 1
    ws.row_dimensions[r].height = 5; r += 1

    wl(r, l03_label, bold=True); r += 1
    wl(r, f"  Opening balance: {'Dr' if ob_l03 < 0 else 'Cr'}"); ws.cell(r,2,abs(ob_l03)); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    mov_label = "  Add: Deficit for year" if l03_movement < 0 else "  Add: Surplus for year"
    wl(r, mov_label); ws.cell(r,2,abs(l03_movement)); ws.cell(r,2).number_format=MONEY; ws.cell(r,2).alignment=_align("right"); r += 1
    wl(r, f"{l03_label} Total", abs(l03) if l03 < 0 else l03, bold=True); r += 1

    if l04 != 0:
        ws.row_dimensions[r].height = 5; r += 1
        wl(r, "Loan from Trustees", l04); r += 1
    if l05 != 0:
        ws.row_dimensions[r].height = 5; r += 1
        wl(r, "Audit Fees Payable", l05); r += 1

    ws.row_dimensions[r].height = 5; r += 1
    total_l(r, "TOTAL", total_fl)
    left_end = r

    # RIGHT SIDE — Assets
    r = start_r
    hdr_r(r, "Assets"); r += 1
    wr(r, "Cash in Hand",                  a01); r += 1
    wr(r, "Cash at Bank — IOB Savings",    a02); r += 1
    wr(r, "Fixed Deposits",                a03); r += 1
    wr(r, "Accrued Interest on FD",        a04); r += 1
    if a05 != 0:
        wr(r, "Advance to Priest — Manikandan", a05); r += 1

    # Pad right side
    while r < left_end:
        ws.row_dimensions[r].height = 14; r += 1

    total_r(left_end, "TOTAL", total_assets)

    _set_width(ws, 1, 38); _set_width(ws, 2, 18)
    _set_width(ws, 3, 36); _set_width(ws, 4, 18)

    return _make_xlsx(wb)


# ══════════════════════════════════════════════════════════════════════════════
# 4. FESTIVAL P&L
# ══════════════════════════════════════════════════════════════════════════════

def festival_pnl_xlsx(fy, rows):
    """
    fy  : str  e.g. '2025-26'
    rows: list of dicts with keys: Fund, 'Income (₹)', 'Expenditure (₹)', 'Surplus / (Deficit) (₹)'
    Returns xlsx bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Festival PnL {fy}"
    year_end = fy.split("-")[1]

    r = _trust_header(ws, "Festival P&L Summary",
                      f"For the year ended 31 March 20{year_end}", fy, 4)

    _write_col_hdr(ws, r,
                   ["Fund", "Income (₹)", "Expenditure (₹)", "Surplus / (Deficit) (₹)"],
                   money_cols=[2, 3, 4])
    r += 1

    total_inc = 0.0
    total_exp = 0.0
    total_sur = 0.0

    for i, row in enumerate(rows):
        inc = float(row.get("Income (₹)", 0) or 0)
        exp = float(row.get("Expenditure (₹)", 0) or 0)
        sur = float(row.get("Surplus / (Deficit) (₹)", 0) or 0)
        total_inc += inc
        total_exp += exp
        total_sur += sur

        bg = _fill(LGREY_HEX) if i % 2 == 0 else _fill(WHITE_HEX)
        sur_color = GREEN_HEX if sur >= 0 else RED_HEX

        c1 = ws.cell(r, 1, row["Fund"])
        c1.font = _font(size=9); c1.fill = bg; c1.border = _border_thin()
        c1.alignment = _align()

        for col, val, color in [(2, inc, None), (3, exp, None), (4, sur, sur_color)]:
            c = ws.cell(r, col, val)
            c.font = _font(size=9, color=color)
            c.fill = bg; c.border = _border_thin()
            c.number_format = MONEY; c.alignment = _align("right")

        ws.row_dimensions[r].height = 16
        r += 1

    # Totals row
    tb = _border_tb(BLUE_HEX, "medium")
    total_sur_color = GREEN_HEX if total_sur >= 0 else RED_HEX

    c = ws.cell(r, 1, "TOTAL")
    c.font = _font(bold=True, size=9); c.border = tb; c.alignment = _align()

    for col, val, color in [
        (2, total_inc, None), (3, total_exp, None), (4, total_sur, total_sur_color)
    ]:
        c = ws.cell(r, col, val)
        c.font = _font(bold=True, size=9, color=color)
        c.number_format = MONEY; c.alignment = _align("right"); c.border = tb

    ws.row_dimensions[r].height = 18

    _set_width(ws, 1, 34); _set_width(ws, 2, 18)
    _set_width(ws, 3, 20); _set_width(ws, 4, 26)

    return _make_xlsx(wb)
